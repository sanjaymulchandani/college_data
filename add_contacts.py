from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Define the XPaths for elements
google_search_box = '//*[@id="APjFqb"]'  # Updated XPath for the search box
phone_text_google = "//span[text()='Phone']"  # Updated XPath for the phone text
phone_number_google = "//span[@data-dtype='d3BN6e']/span"  # Updated XPath for the phone number

wiki_google_link = '//*[@id="rso"]/div[1]/div/div/div/div[1]/div/div/span/a/h3'
wiki_text = "//th[normalize-space()='Director']"

# Initialize the WebDriver
driver = webdriver.Chrome()
url = 'https://www.google.com'  # Add 'https://' to the URL
driver.get(url)
driver.maximize_window()

file_path = 'new_bk.xlsx'
workbook = openpyxl.load_workbook(file_path)
active_sheet = workbook.active

row_count = active_sheet.max_row
print(row_count)
col_count = active_sheet.max_column
print(col_count)

def google_data():
    for i in range(1, row_count + 1):
        college_name = active_sheet.cell(row=i, column=1).value
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, google_search_box)))
        search_box = driver.find_element(By.XPATH, google_search_box)
        search_box.clear()  # Clear the search box before entering a new query
        search_box.send_keys(college_name)  # Adding 'university' to the query
        search_box.submit()  # Submit the search form

        try:
            # Wait for the phone text element to appear and extract the phone number
            phone_text_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, phone_text_google)))
            phone_number_element = driver.find_element(By.XPATH, phone_number_google)
            phone_number = phone_number_element.text
            active_sheet.cell(row=i, column=16).value = phone_number
        except Exception as e:
            print(f"Error for {college_name}: {e}")
            active_sheet.cell(row=i, column=16).value = "NA"

google_data()

# Save the modified workbook
workbook.save(file_path)

# Close the WebDriver
driver.quit()







# Below code is used to fetch director name of college

"""for i in range(1, row_count + 1):
    college_name = active_sheet.cell(row=i, column=1).value
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, google_search_box)))
    search_box = driver.find_element(By.XPATH, google_search_box)
    search_box.clear()  # Clear the search box before entering a new query
    search_box.send_keys(college_name + 'Wikipedia')  # Adding 'university' to the query
    search_box.submit()  # Submit the search form

    try:
        # Wait for the phone text element to appear and extract the phone number
        driver.find_element(By.XPATH, wiki_google_link).click()
        wiki_text_element = driver.find_element((By.XPATH, wiki_text))


    except Exception as e:
        print(f"Error for {college_name}: {e}")
        active_sheet.cell(row=i, column=16).value = "NA"

# Save the modified workbook
workbook.save(file_path)

# Close the WebDriver
driver.quit()
"""